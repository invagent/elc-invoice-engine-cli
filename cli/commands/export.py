"""export 命令组：elc export init / elc export save-mapping / elc export invoice。"""
from __future__ import annotations

import calendar
import json
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import httpx
import typer
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.table import Table

from cli.config import BASE_URL, REQUEST_TIMEOUT
from cli.export.excel_writer import apply_mapping, write_excel
from cli.export.mapping_store import load_mapping, mapping_exists, save_mapping, MAPPINGS_DIR
from cli.export.xml_parser import extract_fields, flatten_xml_fields
from cli.session import get_client

app = typer.Typer(help="发票数据导出")
console = Console()


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _read_template_cols(template_path: str) -> tuple[list[tuple], list[tuple]]:
    """读取模版两个 Sheet 的列名，返回 [(英文, 中文), ...]。"""
    from openpyxl import load_workbook
    wb = load_workbook(template_path, read_only=True)

    def _pairs(ws) -> list[tuple[str, str]]:
        row1 = [c.value or "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        row2 = [c.value or "" for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
        return [(str(a), str(b)) for a, b in zip(row1, row2) if a]

    return _pairs(wb["表头信息"]), _pairs(wb["商品详情"])


def _fetch_source_xml(invoice_id: str, token: str, company_id: str) -> bytes | None:
    """下载 KDUBL Source XML，返回原始字节。"""
    headers = {
        "Authorization": f"Bearer {token}",
        "X-Company-Id":  company_id,
    }
    url = f"{BASE_URL}/v1/invoices/file/{invoice_id}"
    params = {"filetype": "Source", "fileFormat": "xml"}
    try:
        resp = httpx.get(url, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 200:
            ct = resp.headers.get("content-type", "")
            if "xml" in ct or resp.content.strip().startswith(b"<"):
                return resp.content
    except Exception:
        pass
    return None


# invoiceRequests 列表接口内嵌的 invoice 字段（不需要调详情接口即可获取）
_INLINE_INVOICE_FIELDS = {
    "invoiceId", "invoiceNo", "invoiceStatus", "issueStatus",
    "totalAmount", "taxAmount", "taxExclusiveAmount",
    "invoiceTypeCode", "invoiceCurrencyCode", "issueDate",
    "customerName", "customerId", "supplierName", "supplierId",
    "supplierCountryCode",
}


def _needs_detail_api(mapping: dict) -> bool:
    """判断映射表是否需要调 /v2/invoices/{id} 详情接口。"""
    for rule in {**mapping.get("header_mapping", {}), **mapping.get("line_mapping", {})}.values():
        if rule.get("source") == "json":
            path = rule.get("path", "").split(".")[0]
            if path and path not in _INLINE_INVOICE_FIELDS:
                return True
    return False


def _collect_invoices(
    updated_from: str,
    updated_to: str,
    max_batches: int = 500,
) -> list[dict]:
    """分页拉取 INVOICED_SUCCESS 的 invoice 对象列表，内嵌部分字段直接从列表接口取。"""
    client = get_client()
    all_invoices: list[dict] = []
    seen_invoice_ids: set = set()
    cursor = None
    batch = 0

    while batch < max_batches:
        params: dict = {
            "pageSize":    20,
            "updatedFrom": updated_from,
            "updatedTo":   updated_to,
        }
        if cursor is not None:
            params["cursor"] = cursor

        status, body = client.get("/v2/invoice-requests", params=params)
        if status != 200 or body.get("code") not in (None, "0000"):
            console.print(f"[red]查询失败：{body}[/red]")
            break

        data = body.get("data") or {}
        requests = data.get("invoiceRequests", [])
        batch += 1

        for ir in requests:
            if ir.get("invoiceRequestStatus") != "INVOICED_SUCCESS":
                continue
            for inv in ir.get("invoices", []):
                if inv.get("issueStatus") == "SUCCESS":
                    inv_id = inv["invoiceId"]
                    if inv_id not in seen_invoice_ids:
                        seen_invoice_ids.add(inv_id)
                        # 合并申请单级字段到发票对象
                        merged = {**inv, "sourceSystem": ir.get("sourceSystem", ""),
                                  "invoiceRequestId": ir.get("invoiceRequestId", "")}
                        all_invoices.append(merged)

        if not data.get("hasMore"):
            break

        next_cursor = data.get("nextCursor")
        if next_cursor is None or next_cursor == cursor:
            break
        cursor = next_cursor

    console.print(f"  共查询 {batch} 批，找到 {len(all_invoices)} 张不重复发票")
    return all_invoices


def _resolve_template() -> str | None:
    """从缓存目录推断模版路径。"""
    if MAPPINGS_DIR.exists():
        cached = list(MAPPINGS_DIR.glob("*.json"))
        if cached:
            try:
                m = json.loads(cached[0].read_text())
                return m.get("template_path", "")
            except Exception:
                pass
    return None


# ── 命令：init（采集样本数据，输出给 Claude 推断）────────────────────────────

@app.command("init")
def export_init(
    template: Optional[str] = typer.Option(None, "--template", "-t", help="Excel 模版路径，默认使用已缓存的模版"),
    company: Optional[str] = typer.Option(None, "--company", "-c", help="公司 ID"),
    output: Optional[str] = typer.Option(None, "--output", "-o",
                                          help="将样本数据输出到文件（默认打印到终端）"),
):
    """采集模版列名和真实发票样本，输出给 Claude 推断字段映射。

    使用步骤：
      1. 运行此命令，将输出复制给 Claude
      2. Claude 推断映射后生成 JSON
      3. 运行 elc export save-mapping 保存映射
    """
    if template is None:
        template = _resolve_template()
        if not template:
            console.print("[red]未找到已缓存的模版，请通过 --template 指定模版路径[/red]")
            raise typer.Exit(1)
        console.print(f"[dim]使用已缓存模版：{template}[/dim]")

    template = str(Path(template).expanduser().resolve())
    if not Path(template).exists():
        console.print(f"[red]模版文件不存在：{template}[/red]")
        raise typer.Exit(1)

    company = company or os.environ.get("X_COMPANY_ID", "")

    console.print("[cyan]正在读取模版列名...[/cyan]")
    header_cols, line_cols = _read_template_cols(template)

    # 拉一条真实发票样本
    console.print("[cyan]正在获取样本数据...[/cyan]")
    client = get_client()
    status, body = client.get("/v2/invoice-requests", params={"pageSize": 20})
    requests = (body.get("data") or {}).get("invoiceRequests", [])
    sample_ir = next(
        (r for r in requests
         if r.get("invoiceRequestStatus") == "INVOICED_SUCCESS" and r.get("invoices")),
        None,
    )
    if not sample_ir:
        console.print("[red]未找到已开票成功的申请单，无法采集样本。[/red]")
        raise typer.Exit(1)

    inv_id = sample_ir["invoices"][0]["invoiceId"]

    # 同时采集内嵌数据和详情数据，供 AI 推断时参考
    inline_invoice = sample_ir["invoices"][0]
    _, inv_body = client.get(f"/v2/invoices/{inv_id}")
    invoice_json = inv_body.get("data") or inv_body

    # 下载 KDUBL XML
    console.print("[cyan]正在下载 KDUBL XML 样本...[/cyan]")
    token = client.headers.get("Authorization", "").removeprefix("Bearer ")
    xml_bytes = _fetch_source_xml(inv_id, token, company)
    xml_fields_list = flatten_xml_fields(xml_bytes) if xml_bytes else []

    # 构建输出数据，注明内嵌字段范围供 AI 判断是否需要详情接口
    result = {
        "template_path": template,
        "header_columns": [{"en": en, "zh": zh} for en, zh in header_cols],
        "line_columns":   [{"en": en, "zh": zh} for en, zh in line_cols],
        "inline_invoice_fields": sorted(_INLINE_INVOICE_FIELDS),
        "sample_inline_invoice": inline_invoice,
        "sample_invoice_json": invoice_json,
        "sample_xml_fields":   xml_fields_list[:80],
    }

    out_str = json.dumps(result, ensure_ascii=False, indent=2)

    if output:
        Path(output).write_text(out_str)
        console.print(f"[green]✓ 样本数据已保存到：{output}[/green]")
        console.print("请将文件内容提供给 Claude，让其推断字段映射。")
    else:
        console.print("\n[bold yellow]── 样本数据（请复制给 Claude 推断映射）──[/bold yellow]\n")
        console.print(out_str)
        console.print(f"\n[dim]模版路径：{template}[/dim]")


# ── 命令：save-mapping（保存 Claude 推断的映射表）──────────────────────────────

@app.command("save-mapping")
def export_save_mapping(
    template: str = typer.Option(..., "--template", "-t", help="Excel 模版路径"),
    mapping_file: Optional[str] = typer.Option(None, "--file", "-f",
                                                help="包含映射 JSON 的文件路径"),
    mapping_json: Optional[str] = typer.Option(None, "--json", "-j",
                                                help="直接传入映射 JSON 字符串"),
):
    """保存字段映射表到本地缓存（~/.elc/mappings/）。

    映射 JSON 格式：
      {
        "header_mapping": {"*invoiceNo": {"source": "json", "path": "invoiceNo"}, ...},
        "line_mapping":   {"*itemName":  {"source": "json", "path": "itemName"},  ...}
      }
    """
    template = str(Path(template).expanduser().resolve())
    if not Path(template).exists():
        console.print(f"[red]模版文件不存在：{template}[/red]")
        raise typer.Exit(1)

    if mapping_file:
        mapping = json.loads(Path(mapping_file).expanduser().read_text())
    elif mapping_json:
        mapping = json.loads(mapping_json)
    else:
        console.print("[red]需要 --file 或 --json 参数[/red]")
        raise typer.Exit(1)

    if "header_mapping" not in mapping or "line_mapping" not in mapping:
        console.print("[red]映射 JSON 格式错误，需包含 header_mapping 和 line_mapping[/red]")
        raise typer.Exit(1)

    # 分析并记录是否需要详情接口和 XML
    needs_detail = _needs_detail_api(mapping)
    needs_xml = any(
        r.get("source") == "xml"
        for r in {**mapping.get("header_mapping", {}), **mapping.get("line_mapping", {})}.values()
    )
    mapping["_meta"] = {"needs_detail": needs_detail, "needs_xml": needs_xml}

    save_mapping(template, mapping)

    hm = mapping["header_mapping"]
    lm = mapping["line_mapping"]
    json_count = sum(1 for r in {**hm, **lm}.values() if r.get("source") == "json")
    xml_count  = sum(1 for r in {**hm, **lm}.values() if r.get("source") == "xml")
    none_count = sum(1 for r in {**hm, **lm}.values() if r.get("source") == "none")

    console.print(f"[green]✓ 映射表已保存[/green]（JSON字段：{json_count}，XML字段：{xml_count}，无映射：{none_count}）")
    console.print(f"  需要详情接口：{'是' if needs_detail else '否（内嵌字段已满足）'}")
    console.print(f"  需要 XML：{'是' if needs_xml else '否'}")
    console.print(f"  缓存位置：~/.elc/mappings/")


# ── 命令：invoice（导出发票到 Excel）─────────────────────────────────────────

@app.command("invoice")
def export_invoice(
    template: Optional[str] = typer.Option(None, "--template", "-t", help="Excel 模版路径，默认使用已缓存的模版"),
    updated_from: Optional[str] = typer.Option(None, "--from", help="起始日期，如 2026-06-01"),
    updated_to: Optional[str] = typer.Option(None, "--to", help="截止日期，如 2026-06-30"),
    days: Optional[int] = typer.Option(None, "--days", "-d", help="最近 N 天，与 --from/--to 互斥"),
    month: Optional[str] = typer.Option(None, "--month", "-m", help="月份，如 2026-06（兼容旧参数）"),
    company: Optional[str] = typer.Option(None, "--company", "-c", help="公司 ID"),
    out: Optional[str] = typer.Option(None, "--out", "-o", help="输出目录，默认 ~/Downloads"),
    max_batches: int = typer.Option(500, "--max-batches", help="最大查询批次（每批20条）"),
):
    """导出发票数据到 Excel。

    日期范围支持三种方式（优先级：--from/--to > --days > --month > 当月）：
      --from 2026-06-01 --to 2026-06-30
      --days 15              （最近15天）
      --month 2026-06        （整月）
    """
    # 默认输出目录
    if out is None:
        out = str(Path.home() / "Downloads")

    # 推断模版路径
    if template is None:
        template = _resolve_template()
        if not template:
            console.print("[red]未找到已缓存的模版，请通过 --template 指定模版路径[/red]")
            raise typer.Exit(1)
        console.print(f"[dim]使用已缓存模版：{template}[/dim]")

    template = str(Path(template).expanduser().resolve())
    if not Path(template).exists():
        console.print(f"[red]模版文件不存在：{template}[/red]")
        raise typer.Exit(1)

    if not mapping_exists(template):
        console.print("[red]映射表未初始化，请先运行：[/red]")
        console.print(f"  elc export init --template {template}")
        raise typer.Exit(1)

    # 解析日期范围
    today = date.today()
    if updated_from and updated_to:
        pass  # 直接使用
    elif days:
        updated_from = (today - timedelta(days=days)).strftime("%Y-%m-%d")
        updated_to   = today.strftime("%Y-%m-%d")
    elif month:
        d = datetime.strptime(month, "%Y-%m")
        last_day = calendar.monthrange(d.year, d.month)[1]
        updated_from = f"{d.year:04d}-{d.month:02d}-01"
        updated_to   = f"{d.year:04d}-{d.month:02d}-{last_day:02d}"
    else:
        updated_from = today.strftime("%Y-%m-01")
        updated_to   = today.strftime("%Y-%m-%d")

    mapping = load_mapping(template)
    company = company or os.environ.get("X_COMPANY_ID", "")

    # 从映射表元信息决定调用路径
    meta = mapping.get("_meta", {})
    needs_detail = meta.get("needs_detail", True)   # 保守默认：调详情
    needs_xml    = meta.get("needs_xml", False)

    console.print(f"[cyan]日期范围：{updated_from} ~ {updated_to}，公司：{company}[/cyan]")
    console.print(f"[dim]调用路径：{'详情接口' if needs_detail else '内嵌数据（跳过详情接口）'}，XML：{'需要' if needs_xml else '不需要'}[/dim]")

    # ① 收集发票列表（内嵌数据）
    inv_metas = _collect_invoices(updated_from, updated_to, max_batches=max_batches)
    if not inv_metas:
        console.print("[yellow]未找到符合条件的发票，退出。[/yellow]")
        raise typer.Exit(0)
    console.print(f"共找到 [bold]{len(inv_metas)}[/bold] 张发票")

    xml_xpath_map = {
        col: rule["xpath"]
        for col, rule in {**mapping.get("header_mapping", {}), **mapping.get("line_mapping", {})}.items()
        if rule.get("source") == "xml" and rule.get("xpath")
    } if needs_xml else {}

    client = get_client()
    token = client.headers.get("Authorization", "").removeprefix("Bearer ")

    invoices_out: list[dict] = []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("处理发票...", total=len(inv_metas))

        for inline_inv in inv_metas:
            inv_id = inline_inv["invoiceId"]
            progress.update(task, description=f"处理 {inv_id[:16]}...")

            # 按需调详情接口
            if needs_detail:
                _, inv_body = client.get(f"/v2/invoices/{inv_id}")
                invoice_json = inv_body.get("data") or inv_body
                invoice_json["sourceSystem"] = inline_inv.get("sourceSystem", "")
            else:
                invoice_json = inline_inv

            # 按需下载 XML
            xml_fields: dict = {}
            if needs_xml:
                xml_bytes = _fetch_source_xml(inv_id, token, company)
                if xml_bytes:
                    xml_fields = extract_fields(xml_bytes, xml_xpath_map)

            row = apply_mapping(invoice_json, xml_fields, mapping)
            invoices_out.append(row)
            progress.advance(task)

    # ③ 写 Excel
    date_suffix = f"{updated_from}_{updated_to}".replace("-", "")
    out_path = str(Path(out).expanduser() / f"invoice_export_{date_suffix}.xlsx")
    count = write_excel(invoices_out, template, out_path, mapping)

    console.print(f"\n[green]✓ 导出完成：{out_path}[/green]")
    console.print(f"  共导出 [bold]{count}[/bold] 张发票")

