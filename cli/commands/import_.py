"""import 命令组：elc import init / elc import save-mapping / elc import invoice。"""
from __future__ import annotations

import json
import os
import time
import uuid
from pathlib import Path
from typing import Optional

import httpx
import typer
from openpyxl import load_workbook
from rich.console import Console
from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn
from rich.table import Table

from cli.config import BASE_URL, REQUEST_TIMEOUT
from cli.import_.mapping_store import (
    IMPORT_MAPPINGS_DIR, load_mapping, mapping_exists, save_mapping,
)
from cli.import_.xml_builder import build_xml
from cli.session import get_client

app = typer.Typer(help="从 Excel 导入开票数据")
console = Console()


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _read_template_cols(template_path: str) -> tuple[list[tuple], list[tuple]]:
    wb = load_workbook(template_path, read_only=True)

    def _pairs(ws) -> list[tuple[str, str]]:
        row1 = [c.value or "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        row2 = [c.value or "" for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
        return [(str(a), str(b)) for a, b in zip(row1, row2) if a]

    return _pairs(wb["表头信息"]), _pairs(wb["商品详情"])


def _read_excel_data(template_path: str) -> tuple[list[dict], list[dict]]:
    """读取 Excel 数据，返回 (headers, lines)，每个元素是 {col_name: value}。"""
    wb = load_workbook(template_path, read_only=True)

    def _read_sheet(ws) -> list[dict]:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return []
        col_names = [str(c) if c else "" for c in rows[0]]
        return [
            {col_names[i]: row[i] for i in range(len(col_names)) if col_names[i]}
            for row in rows[2:]
            if any(v is not None for v in row)
        ]

    return _read_sheet(wb["表头信息"]), _read_sheet(wb["商品详情"])


def _resolve_template() -> str | None:
    if IMPORT_MAPPINGS_DIR.exists():
        cached = list(IMPORT_MAPPINGS_DIR.glob("*.json"))
        if cached:
            try:
                m = json.loads(cached[0].read_text())
                return m.get("template_path", "")
            except Exception:
                pass
    return None


def _post_invoice_request(
    xml_bytes: bytes,
    source_system: str,
    source_id: str,
    batch_id: str | None,
    client_headers: dict,
) -> tuple[int, dict, str]:
    """提交单个开票申请，返回 (status_code, body, idempotency_key)。"""
    ikey = str(uuid.uuid4())
    headers = {k: v for k, v in client_headers.items() if k.lower() != "content-type"}
    headers.update({
        "X-Request-Id":      uuid.uuid4().hex,
        "X-Idempotency-Key": ikey,
        "X-Timezone":        "UTC+00:00",
    })
    form_data: dict = {"sourceSystem": source_system, "sourceId": source_id}
    if batch_id:
        form_data["batchId"] = batch_id

    resp = httpx.post(
        f"{BASE_URL}/v2/invoice-requests",
        headers=headers,
        files={"payload": ("invoice.xml", xml_bytes, "application/xml")},
        data=form_data,
        timeout=REQUEST_TIMEOUT,
    )
    body = resp.json() if "application/json" in resp.headers.get("content-type", "") else {"raw": resp.text}
    return resp.status_code, body, ikey


def _poll_invoice_status(invoice_request_id: str, client, max_wait: int = 60) -> list[dict]:
    """轮询申请单下的发票状态，最多等待 max_wait 秒。"""
    deadline = time.time() + max_wait
    while time.time() < deadline:
        status, body = client.get(f"/v2/invoice-requests/{invoice_request_id}/invoices")
        if status == 200:
            data = body.get("data") or []
            if isinstance(data, list) and data:
                # 所有发票都有终态则返回
                terminal = {"SUCCESS", "FAILED", "CANCELLED"}
                if all(inv.get("issueStatus") in terminal for inv in data):
                    return data
        time.sleep(3)
    # 超时，返回最后一次查询结果
    _, body = client.get(f"/v2/invoice-requests/{invoice_request_id}/invoices")
    return body.get("data") or []


# ── 命令：init ────────────────────────────────────────────────────────────────

@app.command("init")
def import_init(
    template: Optional[str] = typer.Option(None, "--template", "-t",
                                            help="Excel 模版路径，默认使用已缓存的模版"),
    output: Optional[str] = typer.Option(None, "--output", "-o",
                                          help="将样本数据输出到文件（默认打印到终端）"),
):
    """采集 Excel 模版列名和 UBL XML 字段，输出给 Claude 推断 Excel→XML 字段映射。

    使用步骤：
      1. 运行此命令，将输出提供给 Claude
      2. Claude 推断映射后生成 JSON
      3. 运行 elc import save-mapping 保存映射
    """
    if template is None:
        template = _resolve_template()
        if not template:
            console.print("[red]未找到已缓存的模版，请通过 --template 指定模版路径[/red]")
            raise typer.Exit(1)
        console.print(f"[dim]使用已缓存模版：{template}[/dim]")

    template = str(Path(template).expanduser().resolve())
    if not Path(template).exists():
        console.print(f"[red]模版文件不存在：{template}[/red]")
        raise typer.Exit(1)

    console.print("[cyan]正在读取模版列名...[/cyan]")
    header_cols, line_cols = _read_template_cols(template)

    # UBL XML 字段说明（供 Claude 推断映射时参考）
    ubl_header_fields = {
        "InvoiceTypeCode":          "发票类型，如 380（普通发票）/ 381（贷项）/ 383（借项）",
        "DocumentCurrencyCode":     "发票货币代码，如 EUR / USD / CNY",
        "TaxCurrencyCode":          "计税货币代码（跨境贸易用）",
        "DueDate":                  "到期日 YYYY-MM-DD",
        "TaxPointDate":             "税点日期 YYYY-MM-DD",
        "Note":                     "发票备注",
        "BuyerReference":           "买方联系人/邮箱",
        "BillId":                   "业务单据ID（幂等去重）",
        "SupplierPartyId":          "销方 ID（ERP 内部编码，如 BU-00008）",
        "SupplierPartySchemeID":    "销方 ID 类型，默认 ERP",
        "SupplierName":             "销方公司名",
        "SupplierCountry":          "销方国家代码，如 DE / MY / GB",
        "CustomerPartyId":          "买方 ID（税号或 ERP 编码）",
        "CustomerPartySchemeID":    "买方 ID 类型，如 TAXID / ERP",
        "CustomerName":             "买方公司名",
        "CustomerStreet":           "买方街道",
        "CustomerCity":             "买方城市",
        "CustomerPostalZone":       "买方邮编",
        "CustomerCountry":          "买方国家代码",
        "PayeeAccountNo":           "收款账号",
        "PayeeAccountName":         "收款账户名",
        "FinancialInstitutionBranchId": "银行 SWIFT 代码",
        "PaymentTermsNote":         "付款条款说明",
    }
    ubl_line_fields = {
        "Quantity":             "数量",
        "UnitCode":             "单位代码，如 H87 / KGM / C62",
        "LineExtensionAmount":  "行不含税金额",
        "TaxAmount":            "行税额",
        "ItemCode":             "商品编码",
        "ItemName":             "商品名称",
        "ItemDescription":      "商品描述",
        "TaxCategoryCode":      "税率类别，如 S（标准）/ E（免税）/ Z（零税率）",
        "TaxRate":              "税率，如 21 / 0",
        "UnitPrice":            "单价",
    }

    result = {
        "template_path": template,
        "header_columns": [{"en": en, "zh": zh} for en, zh in header_cols],
        "line_columns":   [{"en": en, "zh": zh} for en, zh in line_cols],
        "ubl_header_fields": ubl_header_fields,
        "ubl_line_fields":   ubl_line_fields,
        "mapping_format": {
            "description": "header_mapping / line_mapping 中每个条目：col 为 Excel 列名（英文），default 为固定默认值（col 和 default 二选一）",
            "example": {
                "header_mapping": {
                    "InvoiceTypeCode":       {"col": "*invoiceType",        "default": None},
                    "DocumentCurrencyCode":  {"col": "*documentCurrencyCode", "default": None},
                    "SupplierPartyId":       {"col": None,                  "default": "BU-00008"},
                    "SupplierPartySchemeID": {"col": None,                  "default": "ERP"},
                    "BillId":               {"col": "*sourceId",            "default": None},
                },
                "line_mapping": {
                    "ItemName":    {"col": "*itemName",  "default": None},
                    "Quantity":    {"col": "*quantity",  "default": None},
                    "UnitCode":    {"col": "*unit",      "default": "H87"},
                    "UnitPrice":   {"col": "*unitPrice", "default": None},
                    "TaxRate":     {"col": "*taxRate",   "default": "0"},
                    "TaxCategoryCode": {"col": "*taxCode", "default": "S"},
                },
            },
        },
    }

    out_str = json.dumps(result, ensure_ascii=False, indent=2)
    if output:
        Path(output).write_text(out_str)
        console.print(f"[green]✓ 样本数据已保存到：{output}[/green]")
    else:
        console.print("\n[bold yellow]── 模版列名 & UBL 字段（提供给 Claude 推断映射）──[/bold yellow]\n")
        console.print(out_str)


# ── 命令：save-mapping ────────────────────────────────────────────────────────

@app.command("save-mapping")
def import_save_mapping(
    template: str = typer.Option(..., "--template", "-t", help="Excel 模版路径"),
    mapping_file: Optional[str] = typer.Option(None, "--file", "-f",
                                                help="包含映射 JSON 的文件路径"),
    mapping_json: Optional[str] = typer.Option(None, "--json", "-j",
                                                help="直接传入映射 JSON 字符串"),
):
    """保存 Excel→XML 字段映射表到本地缓存（~/.elc/import-mappings/）。"""
    template = str(Path(template).expanduser().resolve())
    if not Path(template).exists():
        console.print(f"[red]模版文件不存在：{template}[/red]")
        raise typer.Exit(1)

    if mapping_file:
        mapping = json.loads(Path(mapping_file).expanduser().read_text())
    elif mapping_json:
        mapping = json.loads(mapping_json)
    else:
        console.print("[red]需要 --file 或 --json 参数[/red]")
        raise typer.Exit(1)

    if "header_mapping" not in mapping or "line_mapping" not in mapping:
        console.print("[red]映射 JSON 格式错误，需包含 header_mapping 和 line_mapping[/red]")
        raise typer.Exit(1)

    save_mapping(template, mapping)
    console.print(f"[green]✓ Import 映射表已保存[/green]")
    console.print(f"  缓存位置：~/.elc/import-mappings/")


# ── 命令：invoice ─────────────────────────────────────────────────────────────

@app.command("invoice")
def import_invoice(
    file: str = typer.Option(..., "-f", "--file", help="Excel 数据文件路径"),
    template: Optional[str] = typer.Option(None, "--template", "-t",
                                            help="Excel 模版路径，默认使用已缓存的模版"),
    source_system: str = typer.Option("CLI", "--source-system", help="来源系统标识"),
    batch_id: Optional[str] = typer.Option(None, "--batch-id",
                                            help="批次 ID，不传则自动生成（同批次共用）"),
    poll_timeout: int = typer.Option(60, "--poll-timeout",
                                     help="等待开票结果的最大秒数（每张发票）"),
    out: Optional[str] = typer.Option(None, "--out", "-o",
                                       help="报告输出目录，默认 ~/Downloads"),
):
    """从 Excel 批量导入开票数据，提交申请并输出开票报告。

    需先运行 elc import init 并通过 elc import save-mapping 保存映射。
    """
    file = str(Path(file).expanduser().resolve())
    if not Path(file).exists():
        console.print(f"[red]文件不存在：{file}[/red]")
        raise typer.Exit(1)

    # 推断模版路径（默认与数据文件相同，若有缓存则用缓存）
    if template is None:
        template = _resolve_template() or file
        console.print(f"[dim]使用模版：{template}[/dim]")

    template = str(Path(template).expanduser().resolve())
    if not Path(template).exists():
        console.print(f"[red]模版文件不存在：{template}[/red]")
        raise typer.Exit(1)

    if not mapping_exists(template):
        console.print("[red]Import 映射表未初始化，请先运行：[/red]")
        console.print(f"  elc import init --template {template}")
        raise typer.Exit(1)

    mapping = load_mapping(template)
    out_dir = str(Path(out).expanduser() if out else Path.home() / "Downloads")
    auto_batch_id = batch_id or str(uuid.uuid4())

    console.print("[cyan]正在读取 Excel 数据...[/cyan]")
    headers_data, lines_data = _read_excel_data(file)

    if not headers_data:
        console.print("[red]表头信息 Sheet 无数据[/red]")
        raise typer.Exit(1)

    console.print(f"读取到 [bold]{len(headers_data)}[/bold] 条申请单，"
                  f"[bold]{len(lines_data)}[/bold] 条行项目")

    client = get_client()

    # 按 BillId（或行号）将 lines 分组到对应 header
    hm = mapping.get("header_mapping", {})
    bill_id_col = hm.get("BillId", {}).get("col")

    # 构建 header → lines 的分组
    grouped: list[tuple[dict, list[dict]]] = []
    if bill_id_col:
        # 按 BillId 分组
        groups: dict[str, list] = {}
        for h in headers_data:
            bid = str(h.get(bill_id_col, "")) or f"ROW-{len(groups)+1}"
            groups.setdefault(bid, (h, []))
        lm = mapping.get("line_mapping", {})
        line_bill_col = lm.get("BillId", {}).get("col") or bill_id_col
        for line in lines_data:
            bid = str(line.get(line_bill_col, ""))
            if bid in groups:
                groups[bid][1].append(line)
        grouped = list(groups.values())
    else:
        # 无 BillId：假设一个 header 对应所有 lines
        for h in headers_data:
            grouped.append((h, lines_data))

    console.print(f"共 [bold]{len(grouped)}[/bold] 张发票待提交，批次 ID：{auto_batch_id}")

    results: list[dict] = []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("提交开票...", total=len(grouped))

        for idx, (header, lines) in enumerate(grouped, 1):
            source_id = str(header.get(bill_id_col, "")) if bill_id_col else f"CLI-{uuid.uuid4().hex[:8].upper()}"
            progress.update(task, description=f"[{idx}/{len(grouped)}] 构建 XML {source_id}...")

            # 构建 XML
            try:
                xml_bytes = build_xml(header, lines, mapping)
            except Exception as e:
                results.append({
                    "index": idx, "sourceId": source_id,
                    "submitStatus": "BUILD_FAILED", "error": str(e),
                    "invoiceRequestId": "", "invoices": [],
                })
                progress.advance(task)
                continue

            # 提交
            progress.update(task, description=f"[{idx}/{len(grouped)}] 提交 {source_id}...")
            try:
                status_code, body, ikey = _post_invoice_request(
                    xml_bytes, source_system, source_id, auto_batch_id, client.headers
                )
            except Exception as e:
                results.append({
                    "index": idx, "sourceId": source_id,
                    "submitStatus": "SUBMIT_FAILED", "error": str(e),
                    "invoiceRequestId": "", "invoices": [],
                })
                progress.advance(task)
                continue

            if status_code not in (200, 201):
                results.append({
                    "index": idx, "sourceId": source_id,
                    "submitStatus": "SUBMIT_FAILED",
                    "error": body.get("message") or str(body),
                    "invoiceRequestId": "", "invoices": [],
                })
                progress.advance(task)
                continue

            ir_id = (body.get("data") or {}).get("invoiceRequestId", "")
            results.append({
                "index": idx, "sourceId": source_id,
                "submitStatus": "SUBMITTED",
                "invoiceRequestId": ir_id,
                "invoices": [],
            })
            progress.advance(task)

    # 轮询开票状态
    submitted = [r for r in results if r["submitStatus"] == "SUBMITTED" and r["invoiceRequestId"]]
    if submitted:
        console.print(f"\n[cyan]正在等待 {len(submitted)} 张发票开票结果（最多 {poll_timeout}s/张）...[/cyan]")
        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TaskProgressColumn(), console=console) as progress:
            task = progress.add_task("轮询状态...", total=len(submitted))
            for r in submitted:
                progress.update(task, description=f"轮询 {r['invoiceRequestId'][:16]}...")
                invoices = _poll_invoice_status(r["invoiceRequestId"], client, poll_timeout)
                r["invoices"] = invoices
                progress.advance(task)

    # 输出报告
    _print_report(results)
    _save_report(results, out_dir)


def _print_report(results: list[dict]) -> None:
    console.print("\n[bold cyan]── 开票导入报告 ──[/bold cyan]\n")

    table = Table(show_header=True, header_style="bold")
    table.add_column("#",           width=4)
    table.add_column("sourceId",    width=20)
    table.add_column("提交状态",     width=14)
    table.add_column("申请单ID",     width=36)
    table.add_column("发票号",       width=20)
    table.add_column("开票状态",     width=12)

    for r in results:
        submit_style = "green" if r["submitStatus"] == "SUBMITTED" else "red"
        invoices = r.get("invoices") or []
        if invoices:
            for inv in invoices:
                issue_status = inv.get("issueStatus", "-")
                inv_no = inv.get("invoiceNo", "-")
                style = "green" if issue_status == "SUCCESS" else "red"
                table.add_row(
                    str(r["index"]),
                    r["sourceId"],
                    f"[{submit_style}]{r['submitStatus']}[/{submit_style}]",
                    r["invoiceRequestId"],
                    inv_no,
                    f"[{style}]{issue_status}[/{style}]",
                )
        else:
            error = r.get("error", "")
            table.add_row(
                str(r["index"]),
                r["sourceId"],
                f"[{submit_style}]{r['submitStatus']}[/{submit_style}]",
                r.get("invoiceRequestId", ""),
                error[:30] if error else "-",
                "-",
            )

    console.print(table)

    total   = len(results)
    success = sum(1 for r in results
                  for inv in (r.get("invoices") or [])
                  if inv.get("issueStatus") == "SUCCESS")
    failed  = sum(1 for r in results if r["submitStatus"] != "SUBMITTED")
    pending = total - success - failed

    console.print(f"\n合计：{total} 张  |  "
                  f"[green]成功：{success}[/green]  |  "
                  f"[red]失败：{failed}[/red]  |  "
                  f"[yellow]待处理：{pending}[/yellow]")


def _save_report(results: list[dict], out_dir: str) -> None:
    import json as _json
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path(out_dir).expanduser() / f"import_report_{ts}.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_json.dumps(results, ensure_ascii=False, indent=2))
    console.print(f"\n[dim]报告已保存：{path}[/dim]")
