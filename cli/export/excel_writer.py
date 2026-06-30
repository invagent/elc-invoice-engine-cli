"""Excel 模版填写：将映射后的发票数据写入 openpyxl workbook。"""
from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook


def write_excel(
    invoices: list[dict],
    template_path: str,
    output_path: str,
    mapping: dict,
) -> int:
    """
    将发票数据按映射表写入 Excel 模版。

    invoices: 每个元素包含：
      - _header: {col_name: value}  表头信息字段
      - _lines:  [{col_name: value}, ...]  行项目字段
    mapping: load_mapping() 返回的映射表
    返回写入的发票数量。
    """
    wb = load_workbook(template_path)
    header_ws = wb["表头信息"]
    line_ws   = wb["商品详情"]

    # 读取模版列顺序（第1行英文列名）
    header_cols = [c.value for c in header_ws[1]]
    line_cols   = [c.value for c in line_ws[1]]

    header_row = 3  # 第1行英文列名，第2行中文说明，第3行起写数据
    line_row   = 3

    for inv in invoices:
        flat   = inv.get("_header", {})
        lines  = inv.get("_lines", [])

        # 写表头信息行
        for col_idx, col_name in enumerate(header_cols, 1):
            if col_name:
                header_ws.cell(row=header_row, column=col_idx, value=flat.get(col_name))
        header_row += 1

        # 写商品详情行
        for line in lines:
            for col_idx, col_name in enumerate(line_cols, 1):
                if col_name:
                    line_ws.cell(row=line_row, column=col_idx, value=line.get(col_name))
            line_row += 1

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(invoices)


def apply_mapping(invoice_json: dict, xml_fields: dict, mapping: dict) -> dict:
    """
    将一条发票的 JSON + XML 字段按映射表合并为 {col_name: value} 扁平字典。
    返回 {"_header": {...}, "_lines": [{...}, ...]}
    """
    header_map = mapping.get("header_mapping", {})
    line_map   = mapping.get("line_mapping", {})

    # 表头字段
    header_flat: dict = {}
    for col, rule in header_map.items():
        source = rule.get("source")
        if source == "json":
            header_flat[col] = _get_nested(invoice_json, rule.get("path", ""))
        elif source == "xml":
            header_flat[col] = xml_fields.get(col)
        # source == "none" 留空

    # 行项目字段
    lines_out: list[dict] = []
    raw_lines = invoice_json.get("lines", [])
    for raw_line in raw_lines:
        line_flat: dict = {}
        for col, rule in line_map.items():
            source = rule.get("source")
            if source == "json":
                line_flat[col] = _get_nested(raw_line, rule.get("path", ""))
            elif source == "xml":
                line_flat[col] = xml_fields.get(col)
        # invoiceNo 注入到每行
        line_flat["*invoiceNo"] = invoice_json.get("invoiceNo")
        lines_out.append(line_flat)

    return {"_header": header_flat, "_lines": lines_out}


def _get_nested(obj: dict, path: str) -> object:
    """支持点分隔路径，如 'taxInfo.0.taxRate'。"""
    if not path:
        return None
    parts = path.split(".")
    cur = obj
    for p in parts:
        if cur is None:
            return None
        if isinstance(cur, list):
            try:
                cur = cur[int(p)]
            except (ValueError, IndexError):
                return None
        elif isinstance(cur, dict):
            cur = cur.get(p)
        else:
            return None
    return cur
